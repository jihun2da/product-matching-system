# -*- coding: utf-8 -*-
"""
perfect_matcher.py
설정 분리형 구조
- config.py에서 모든 설정을 불러와 사용
- 고속 매칭(블로킹+벡터화) + 상품명/사이즈 캐논화 + 색상 정규화
- 엑셀 색상 적용/가격 비교/결과 저장 포함
"""

import re
import time
import logging
import unicodedata
from dataclasses import dataclass
from typing import List, Tuple, Dict, Optional, Any, Set
from collections import defaultdict

import pandas as pd
from rapidfuzz import fuzz, process
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ---- 설정 불러오기 ----
from config import (
    DEFAULT_PATHS,
    NAME_SYNONYM_GROUPS,
    SIZE_SYNONYMS,
    COLOR_ALIASES,
    EXCEL_COLORS,
    MATCHING,
    LOGGING,
    PERFORMANCE,
)

# ----------------------------------
# 로깅 세팅
# ----------------------------------
logging.basicConfig(
    level=getattr(logging, LOGGING.get("level", "INFO")),
    format="%(asctime)s | %(levelname)s | %(message)s"
)
logger = logging.getLogger("matcher")

# ----------------------------------
# 결과 구조
# ----------------------------------
@dataclass
class MatchResult:
    receipt_idx: int
    matched_idx: int
    confidence_score: float
    match_details: Dict[str, Any]

# ----------------------------------
# 유틸
# ----------------------------------
def make_fill(argb_hex: str) -> PatternFill:
    """ARGB 8자리 HEX로 PatternFill 생성"""
    code = argb_hex.upper()
    if len(code) != 8:
        raise ValueError(f"엑셀 색상 코드는 ARGB 8자리여야 합니다: '{argb_hex}'")
    return PatternFill(start_color=code, end_color=code, fill_type="solid")

# ----------------------------------
# 매칭 엔진
# ----------------------------------
class PerfectMatcher:
    def __init__(self):
        # config에서 읽은 값 보관
        self.name_score_cutoff: int = MATCHING.get("name_score_cutoff", 70)
        w = MATCHING.get("weights", {"brand": 0.25, "name": 0.35, "color": 0.25, "size": 0.15})
        self.weight_brand = float(w.get("brand", 0.25))
        self.weight_name  = float(w.get("name",  0.35))
        self.weight_color = float(w.get("color", 0.25))
        self.weight_size  = float(w.get("size",  0.15))

        self.progress_update_every = PERFORMANCE.get("progress_update_every", 50)

    # ---------- 정규화 유틸 ----------
    def normalize_text_ultra(self, s: str) -> str:
        """브랜드/상품명용 강력 정규화 (소문자화, 기호 제거, 공백 정리)"""
        if not isinstance(s, str):
            s = "" if pd.isna(s) else str(s)
        s = s.strip()
        s = unicodedata.normalize("NFKC", s)
        s = s.lower()
        s = re.sub(r"\s+", " ", s)
        # 한/영/숫자 외는 공백으로 치환 → 붙은 약어도 공백 분리됨 (예: 볼mtm → '볼 mtm')
        s = re.sub(r"[^0-9a-z가-힣]+", " ", s)
        s = re.sub(r"\s+", " ", s).strip()
        return s

    def normalize_color_ultra(self, s: str) -> str:
        """색상 전용 정규화(설정의 COLOR_ALIASES 반영)"""
        s = self.normalize_text_ultra(s)
        tokens = s.split()
        mapped = [COLOR_ALIASES.get(t, t) for t in tokens]
        return " ".join(mapped)

    # ====== 사이즈/상품명 캐논화 유틸 ======
    def _normalize_size_token(self, tok: str) -> str:
        """사이즈 토큰을 캐논 라벨로 통일. 예: '2xl' -> 'xxl'"""
        t = tok.lower()
        for canon, variants in SIZE_SYNONYMS.items():
            if t == canon or t in variants:
                return canon
        return t

    def apply_bidir_synonyms_text(self, normalized_text: str) -> str:
        """
        normalize_text_ultra() 이후의 문자열에
        NAME_SYNONYM_GROUPS를 적용해 캐논 라벨로 통일.
        - 대/소문자 무시
        - 단어 안에 섞여 있어도(포함치환) 치환
        """
        if not normalized_text:
            return normalized_text
        s = normalized_text
        for canon, variants in NAME_SYNONYM_GROUPS.items():
            for v in variants:
                s = re.sub(re.escape(v), canon, s, flags=re.IGNORECASE)
        return s

    def _extract_parentheses_values(self, size_token: str) -> Set[str]:
        """
        괄호가 포함된 사이즈에서 괄호 안팎 값을 각각 추출
        예: '7(M)' -> {'7', 'M'}, 'M(9)' -> {'M', '9'}, 'L' -> {'L'}
        """
        values = set()
        
        # 괄호 패턴 매칭: 괄호 밖(괄호 안) 형태
        parentheses_match = re.match(r'^([^()]+)\(([^()]+)\)$', size_token)
        if parentheses_match:
            outside = parentheses_match.group(1).strip()
            inside = parentheses_match.group(2).strip()
            if outside:
                values.add(outside)
            if inside:
                values.add(inside)
        else:
            # 괄호가 없는 경우 그대로 추가
            if size_token:
                values.add(size_token)
        
        return values

    def extract_size_parts_ultra(self, s: str) -> Set[str]:
        """
        사이즈를 토큰셋으로 추출(S~XL, 5호, 110, free 등)
        - 토큰 캐논화: 2xl → xxl 등
        - 괄호 안팎 값 분리: 7(M) -> {7, M}
        """
        if not isinstance(s, str):
            s = "" if pd.isna(s) else str(s)
        s = unicodedata.normalize("NFKC", s)
        s = s.replace("~", " ~ ")
        s = re.sub(r"[,/|]+", " ", s)
        s = re.sub(r"\s+", " ", s).strip().lower()

        tokens = set()
        for tok in s.split():
            # 기본 정리 (특수문자 제거는 괄호 처리 후에)
            tok = tok.strip()
            if not tok:
                continue
            
            # 괄호 안팎 값 추출
            parentheses_values = self._extract_parentheses_values(tok)
            
            for value in parentheses_values:
                # 각 값에 대해 정리 및 캐논화 적용
                cleaned_value = re.sub(r"[^0-9a-z가-힣]+", "", value)
                if cleaned_value:
                    normalized_value = self._normalize_size_token(cleaned_value)  # ★ 사이즈 캐논화
                    tokens.add(normalized_value)
        
        return tokens

    # ---------- 퍼지 ----------
    def fast_fuzzy_match(self, a: str, b: str) -> float:
        """상품명 매칭용 경량 퍼지(부분비교)"""
        if not a or not b:
            return 0.0
        return float(fuzz.partial_ratio(a, b))

    # ---------- 신뢰도 계산(조기종료 적용) ----------
    def calculate_match_confidence_perfect(self, receipt_row: pd.Series, matched_row: pd.Series):
        try:
            # 정규화 컬럼 우선 사용
            rb = receipt_row.get("r_brand_n") or self.normalize_text_ultra(receipt_row.get("브랜드", ""))
            rn = receipt_row.get("r_name_n")  or self.normalize_text_ultra(receipt_row.get("상품명", ""))
            rc = receipt_row.get("r_color_n") or self.normalize_color_ultra(receipt_row.get("색상", ""))
            rs = receipt_row.get("r_sizes_s")
            if rs is None:
                rs = self.extract_size_parts_ultra(receipt_row.get("사이즈", ""))

            mb = matched_row.get("m_brand_n") or self.normalize_text_ultra(matched_row.get("브랜드", ""))
            mn = matched_row.get("m_name_n")  or self.normalize_text_ultra(matched_row.get("상품명", ""))
            mc = matched_row.get("m_color_n") or self.normalize_color_ultra(matched_row.get("색상", ""))
            ms = matched_row.get("m_sizes_s")
            if ms is None:
                ms = self.extract_size_parts_ultra(matched_row.get("사이즈", ""))

            # 1) 저비용 확정 체크
            if not rs or not ms or len(rs.intersection(ms)) == 0:
                return False, 0.0, {}
            if rb != mb:
                return False, 0.0, {}
            if rc != mc:
                return False, 0.0, {}

            # 2) 상품명 퍼지 (브랜드/색상/사이즈는 확정 통과)
            name_score = self.fast_fuzzy_match(rn, mn)
            if name_score < self.name_score_cutoff:
                return False, 0.0, {}

            # 최종 가중합
            brand_score = 100.0
            color_score = 100.0
            size_score  = 100.0
            total_conf = (
                brand_score * self.weight_brand +
                name_score  * self.weight_name  +
                color_score * self.weight_color +
                size_score  * self.weight_size
            )

            details = {
                "brand_match": True, "brand_score": brand_score,
                "name_match": True,  "name_score": name_score,
                "color_match": True, "color_score": color_score,
                "size_match": True,  "size_score": size_score
            }
            return True, float(total_conf), details

        except Exception as e:
            logger.exception(f"매칭 신뢰도 계산 오류: {e}")
            return False, 0.0, {}

    # ---------- 사전정규화/인덱스 ----------
    def _precompute_normalized_columns(self, df: pd.DataFrame, prefix: str) -> pd.DataFrame:
        brand = df.get("브랜드", "").fillna("")
        name  = df.get("상품명", "").fillna("")
        color = df.get("색상", "").fillna("")
        size  = df.get("사이즈", "").fillna("")

        df[f"{prefix}_brand_n"] = [self.normalize_text_ultra(x) for x in brand]

        # 1차 정규화 → 2차 동의어 캐논화(상품명)
        name_n = [self.normalize_text_ultra(x) for x in name]
        name_n = [self.apply_bidir_synonyms_text(x) for x in name_n]
        df[f"{prefix}_name_n"] = name_n

        df[f"{prefix}_color_n"] = [self.normalize_color_ultra(x) for x in color]
        df[f"{prefix}_sizes_s"] = [self.extract_size_parts_ultra(x) for x in size]
        return df

    def _build_candidate_index(self, matched_df: pd.DataFrame):
        idx_bcs, idx_bc, idx_b = defaultdict(list), defaultdict(list), defaultdict(list)
        for midx, row in matched_df.iterrows():
            b = row["m_brand_n"]; c = row["m_color_n"]; sizes = row["m_sizes_s"]
            idx_b[(b,)].append(midx)
            idx_bc[(b, c)].append(midx)
            if sizes:
                for s in sizes:
                    idx_bcs[(b, c, s)].append(midx)
        return idx_bcs, idx_bc, idx_b

    # ---------- 빠른 매칭기 ----------
    def fuzzy_match_orders_blocked(self, receipt_df: pd.DataFrame, matched_df: pd.DataFrame) -> List[MatchResult]:
        logger.info(f"[FAST] 블로킹+벡터화 매칭 시작: R={len(receipt_df)}, M={len(matched_df)}")

        receipt_df = self._precompute_normalized_columns(receipt_df.copy(), "r")
        matched_df = self._precompute_normalized_columns(matched_df.copy(), "m")

        idx_bcs, idx_bc, idx_b = self._build_candidate_index(matched_df)

        matched_results: List[MatchResult] = []
        receipt_qty = {idx: int(receipt_df.at[idx, "수량"]) if "수량" in receipt_df.columns else 1
                       for idx in receipt_df.index}
        matched_qty = {idx: int(matched_df.at[idx, "수량"]) if "수량" in matched_df.columns else 1
                       for idx in matched_df.index}

        update_every = max(self.progress_update_every, len(receipt_df)//50 or 1)

        with tqdm(total=len(receipt_df), desc="⚡ FAST 매칭", unit="건",
                  bar_format="{l_bar}{bar}| {n_fmt}/{total_fmt} [{elapsed}<{remaining}, {rate_fmt}]") as pbar:

            for r_idx, r_row in receipt_df.iterrows():
                remain = receipt_qty.get(r_idx, 1)
                if remain <= 0:
                    if (len(matched_results) % update_every) == 0:
                        pbar.set_postfix_str(f"matches={len(matched_results)}")
                    pbar.update(1)
                    continue

                b = r_row["r_brand_n"]; c = r_row["r_color_n"]; sizes = r_row["r_sizes_s"]
                candidate_ids: List[int] = []

                # 1순위: (b,c,사이즈)
                if sizes:
                    for s in sizes:
                        candidate_ids.extend(idx_bcs.get((b, c, s), []))
                # 2순위: (b,c)
                if not candidate_ids:
                    candidate_ids.extend(idx_bc.get((b, c), []))
                # 3순위: (b)
                if not candidate_ids:
                    candidate_ids.extend(idx_b.get((b,), []))

                # 재고 있는 것만
                candidate_ids = [mid for mid in candidate_ids if matched_qty.get(mid, 0) > 0]
                if not candidate_ids:
                    if (len(matched_results) % update_every) == 0:
                        pbar.set_postfix_str(f"matches={len(matched_results)}")
                    pbar.update(1)
                    continue

                # 퍼지매칭(상품명만) - 벡터화 호출
                r_name = r_row["r_name_n"]
                cand_names = [matched_df.at[mid, "m_name_n"] for mid in candidate_ids]

                scored = process.extract(
                    r_name,
                    cand_names,
                    scorer=fuzz.partial_ratio,
                    score_cutoff=self.name_score_cutoff,
                    limit=None
                )

                if not scored:
                    if (len(matched_results) % update_every) == 0:
                        pbar.set_postfix_str(f"matches={len(matched_results)}")
                    pbar.update(1)
                    continue

                # 신뢰도 평가 + 정렬
                results: List[Tuple[int, float, Dict[str, Any]]] = []
                for _, _, local_idx in scored:
                    m_idx = candidate_ids[local_idx]
                    ok, conf, details = self.calculate_match_confidence_perfect(
                        receipt_row=r_row, matched_row=matched_df.loc[m_idx]
                    )
                    if ok:
                        results.append((m_idx, conf, details))

                if results:
                    results.sort(key=lambda x: x[1], reverse=True)

                    for m_idx, conf, details in results:
                        if remain <= 0:
                            break
                        if matched_qty[m_idx] <= 0:
                            continue

                        match_qty = min(remain, matched_qty[m_idx])
                        for _ in range(int(match_qty)):
                            matched_results.append(
                                MatchResult(receipt_idx=r_idx, matched_idx=m_idx,
                                            confidence_score=conf, match_details=details)
                            )
                        receipt_qty[r_idx] -= match_qty
                        matched_qty[m_idx] -= match_qty
                        remain -= match_qty

                if (len(matched_results) % update_every) == 0:
                    pbar.set_postfix_str(f"matches={len(matched_results)}")
                pbar.update(1)

        logger.info(f"[FAST] 완료: {len(matched_results)}건")
        return matched_results

    # ---------- 느린 매칭기(레퍼런스) ----------
    def fuzzy_match_orders_perfect_sequential(self, receipt_df: pd.DataFrame, matched_df: pd.DataFrame) -> List[MatchResult]:
        logger.info(f"[SLOW] 순차 퍼지 매칭 시작: R={len(receipt_df)}, M={len(matched_df)}")
        matched_results: List[MatchResult] = []
        receipt_qty = {idx: int(receipt_df.at[idx, "수량"]) if "수량" in receipt_df.columns else 1
                       for idx in receipt_df.index}
        matched_qty = {idx: int(matched_df.at[idx, "수량"]) if "수량" in matched_df.columns else 1
                       for idx in matched_df.index}

        with tqdm(total=len(receipt_df), desc="🐢 SLOW 매칭", unit="건") as pbar:
            for r_idx, r_row in receipt_df.iterrows():
                remain = receipt_qty.get(r_idx, 1)
                if remain <= 0:
                    pbar.update(1)
                    continue

                candidates = []
                for m_idx, m_row in matched_df.iterrows():
                    ok, conf, details = self.calculate_match_confidence_perfect(r_row, m_row)
                    if ok and matched_qty.get(m_idx, 0) > 0:
                        candidates.append((m_idx, conf, details))

                candidates.sort(key=lambda x: x[1], reverse=True)

                for m_idx, conf, details in candidates:
                    if remain <= 0:
                        break
                    if matched_qty[m_idx] <= 0:
                        continue

                    match_qty = min(remain, matched_qty[m_idx])
                    for _ in range(int(match_qty)):
                        matched_results.append(
                            MatchResult(receipt_idx=r_idx, matched_idx=m_idx,
                                        confidence_score=conf, match_details=details)
                        )
                    receipt_qty[r_idx] -= match_qty
                    matched_qty[m_idx] -= match_qty
                    remain -= match_qty

                pbar.update(1)

        logger.info(f"[SLOW] 완료: {len(matched_results)}건")
        return matched_results


# ----------------------------------
# 색상 적용 + 파일 저장 포함한 솔루션
# ----------------------------------
class PerfectSolutionMatcher(PerfectMatcher):

    def apply_colors_with_quantity_enhanced(
        self,
        receipt_path: str,
        matched_path: str,
        output_receipt: str,
        output_matched: str,
        match_results: List[MatchResult],
        receipt_df: pd.DataFrame,
        matched_df: pd.DataFrame,
    ) -> None:
        """
        매칭 쌍 기준으로
        - Receipt 워크북: A~E 컬러 채우기, F(금액) 가격 비교 시 빨간색 표시
        - Matched  워크북: H~N 컬러 채우기, N(14열) 비교 표시
        - G열 값이 TRUE면 시안, 아니면 노랑
        - H~N 범위 초기화 및 매칭 정보 기록은 _update_match_info 에서 수행
        """
        # 워크북 로드
        receipt_wb = load_workbook(receipt_path)
        matched_wb = load_workbook(matched_path)

        receipt_ws = receipt_wb.active
        matched_ws = matched_wb.active

        # 채우기 색 (config에서 불러와 생성)
        cyan_fill   = make_fill(EXCEL_COLORS["cyan"])
        yellow_fill = make_fill(EXCEL_COLORS["yellow"])
        red_fill    = make_fill(EXCEL_COLORS["red"])

        # 매칭 쌍 구성 (수량만큼 중복 포함)
        matched_pairs: List[Tuple[int, int]] = [(r.receipt_idx, r.matched_idx) for r in match_results]

        # 색상 적용
        with tqdm(
            matched_pairs,
            desc="🎨 색상 적용",
            unit="쌍",
            bar_format="{l_bar}{bar}| {n_fmt}/{total_fmt} [{elapsed}<{remaining}, {rate_fmt}]"
        ) as pbar:
            for receipt_idx, matched_idx in pbar:
                try:
                    # G열(7) 값으로 구분
                    g_value = receipt_ws.cell(row=receipt_idx + 2, column=7).value
                    fill_color = cyan_fill if str(g_value).strip().upper() == "TRUE" else yellow_fill

                    # Receipt 파일 색상: A(1)~E(5)
                    for col in range(1, 6):
                        receipt_ws.cell(row=receipt_idx + 2, column=col).fill = fill_color

                    # Matched 파일 색상: H(8)~N(14)
                    for col in range(8, 14 + 1):
                        matched_ws.cell(row=matched_idx + 2, column=col).fill = fill_color

                    # 가격 비교
                    try:
                        receipt_price = float(receipt_df.loc[receipt_idx, "금액"])
                        matched_price = float(matched_df.loc[matched_idx, "도매가"]) * \
                                        float(receipt_df.loc[receipt_idx, "수량"])
                        if receipt_price != matched_price:
                            receipt_ws.cell(row=receipt_idx + 2, column=6).fill = red_fill   # F열
                            matched_ws.cell(row=matched_idx + 2, column=14).fill = red_fill # N열(14)
                        else:
                            receipt_ws.cell(row=receipt_idx + 2, column=6).fill = fill_color
                            matched_ws.cell(row=matched_idx + 2, column=14).fill = fill_color
                    except Exception as pe:
                        logger.warning(f"가격 비교 스킵 (행 r{receipt_idx} m{matched_idx}): {pe}")

                except Exception as e:
                    logger.error(f"색상 적용 오류 (행 {receipt_idx}-{matched_idx}): {e}")
                    continue

        # H~N 초기화 + 매칭정보 업데이트 (Receipt 워크북)
        self._update_match_info(receipt_ws, matched_pairs, matched_df)

        # 저장
        receipt_wb.save(output_receipt)
        matched_wb.save(output_matched)
        logger.info(f"결과 파일 저장 완료: {output_receipt}, {output_matched}")

    def _update_match_info(self, receipt_ws, matched_pairs: List[Tuple[int, int]], matched_df: pd.DataFrame) -> None:
        """
        H열~N열 초기화 및 매칭 정보 업데이트
        - H열: "상품명 / 사이즈"
        - I열~: 매칭된 기준데이터의 행번호(actual_row_number) 누적 기입
        """
        try:
            # 초기화
            for row in range(2, receipt_ws.max_row + 1):
                for col in range(8, 15):  # H(8) ~ N(14)
                    receipt_ws.cell(row=row, column=col).value = None

            match_count: Dict[int, int] = defaultdict(int)

            for receipt_idx, matched_idx in matched_pairs:
                row_excel = receipt_idx + 2

                # 기준데이터 요약 텍스트
                matched_product_name = matched_df.loc[matched_idx, "상품명"] if "상품명" in matched_df.columns else ""
                matched_size = matched_df.loc[matched_idx, "사이즈"] if "사이즈" in matched_df.columns else ""
                matched_info = f"{matched_product_name} / {matched_size}".strip(" /")

                current_col_offset = match_count[receipt_idx]
                target_col = 9 + current_col_offset  # I열(9)부터 오른쪽으로

                if current_col_offset == 0:
                    receipt_ws.cell(row=row_excel, column=8).value = matched_info  # H열

                actual_row_number = matched_idx + 2  # 기준 파일의 실제 엑셀 행번호
                if target_col <= 14:
                    receipt_ws.cell(row=row_excel, column=target_col).value = actual_row_number

                match_count[receipt_idx] += 1

        except Exception as e:
            logger.error(f"매칭 정보 업데이트 오류: {e}")

    # ---------- 엔드-투-엔드 ----------
    def process_excel_perfect_solution(
        self,
        receipt_path: str,
        matched_path: str,
        output_receipt: str,
        output_matched: str,
        use_fast: bool = True
    ) -> Tuple[int, List[MatchResult]]:
        """
        파일 읽기 → 매칭(빠른/느린) → 색상 적용 & 저장 → 리포트 로그
        return: (매칭건수, match_results)
        """
        start_time = time.time()
        logger.info("🎯 완벽한 해결책 처리 시작")

        # 1) 파일 읽기
        read_start = time.time()
        logger.info("📖 파일 읽기 중...")
        receipt_df = pd.read_excel(receipt_path, engine="openpyxl")
        matched_df = pd.read_excel(matched_path, engine="openpyxl")
        read_time = time.time() - read_start
        logger.info(f"📖 파일 읽기 완료 - Receipt: {len(receipt_df)}행, Matched: {len(matched_df)}행 ({read_time:.2f}초)")

        # 2) 매칭
        match_start = time.time()
        if use_fast:
            match_results = self.fuzzy_match_orders_blocked(receipt_df, matched_df)
        else:
            match_results = self.fuzzy_match_orders_perfect_sequential(receipt_df, matched_df)
        match_time = time.time() - match_start

        # 3) 색상 적용 + 저장
        color_start = time.time()
        self.apply_colors_with_quantity_enhanced(
            receipt_path, matched_path, output_receipt, output_matched,
            match_results, receipt_df, matched_df
        )
        color_time = time.time() - color_start

        total_time = time.time() - start_time
        logger.info("🎯 완벽한 해결책 완료:")
        logger.info(f"  ├─ 총 소요시간: {total_time:.2f}초")
        logger.info(f"  ├─ 파일 읽기: {read_time:.2f}초")
        logger.info(f"  ├─ 매칭 처리: {match_time:.2f}초")
        logger.info(f"  └─ 색상 적용: {color_time:.2f}초")
        logger.info(f"  📊 매칭 건수: {len(match_results)}건")

        return len(match_results), match_results

    def generate_match_report(self, match_results: List[MatchResult]) -> Dict[str, Any]:
        if not match_results:
            return {"total_matches": 0, "average_confidence": 0, "match_distribution": {}}

        total = len(match_results)
        avg = sum(mr.confidence_score for mr in match_results) / total if total else 0.0
        high = sum(1 for mr in match_results if mr.confidence_score >= 90)
        mid  = sum(1 for mr in match_results if 70 <= mr.confidence_score < 90)
        low  = sum(1 for mr in match_results if mr.confidence_score < 70)

        report = {
            "total_matches": total,
            "average_confidence": avg,
            "match_distribution": {
                "high_confidence": high,
                "medium_confidence": mid,
                "low_confidence": low
            }
        }
        logger.info(f"매칭 리포트: {report}")
        return report


# ----------------------------------
# 메인 실행 예시
# ----------------------------------
def main():
    print("🏆 RapidFuzz 완벽해결책 (설정 분리형)")
    print("🎯 고속 매칭 + 동의어/약어 캐논화 + 엑셀 색상 적용/저장")
    print("=" * 70)

    matcher = PerfectSolutionMatcher()

    receipt_path = DEFAULT_PATHS["receipt"]
    matched_path = DEFAULT_PATHS["matched"]
    output_receipt = DEFAULT_PATHS["output_receipt"]
    output_matched = DEFAULT_PATHS["output_matched"]

    matched_count, match_results = matcher.process_excel_perfect_solution(
        receipt_path, matched_path, output_receipt, output_matched, use_fast=True
    )

    report = matcher.generate_match_report(match_results)

    print("\n" + "=" * 70)
    print("🏆 결과 요약")
    print("=" * 70)
    print(f"✅ 총 매칭 건수: {matched_count:,}건")
    print(f"📊 평균 신뢰도: {report['average_confidence']:.2f}%")
    print(f"📈 고신뢰도: {report['match_distribution']['high_confidence']:,}건")
    print(f"📉 중신뢰도: {report['match_distribution']['medium_confidence']:,}건")
    print(f"📉 저신뢰도: {report['match_distribution']['low_confidence']:,}건")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        logger.error(f"프로그램 실행 중 오류: {e}")
        print(f"❌ 오류 발생: {e}")
